#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
電商訂單清理工具
Step 1: 刪除空白列
Step 2: 子項目補父訂單（含格式）
Step 3: 商品編號智慧校正
Step 4: 刪除 AE 欄為「[促銷贈品]免運費」的列

用法:
  python ec_clean.py <來源訂單.xlsx> [輸出檔名.xlsx]

若未指定輸出檔名，自動產生「<原始檔名>_已整理.xlsx」
"""
import openpyxl, os, shutil, sys, re, io


def clean(val):
    """清理數值：去除 =、引號等前綴"""
    if val is None:
        return ''
    s = str(val).strip()
    if s.startswith('='):
        s = s[1:]
    s = s.strip('"')
    s = s.strip("'")
    return s.strip()


def is_pure_digits(s):
    """判斷是否為純數字字串"""
    return bool(re.match(r'^\d+$', s))


def main():
    # Fix stdout encoding for Windows
    if sys.platform == 'win32':
        sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')

    if len(sys.argv) < 2:
        print("用法: python ec_clean.py <來源訂單.xlsx> [輸出檔名.xlsx]")
        print()
        print("範例:")
        print("  python ec_clean.py 訂單.xlsx")
        print("  python ec_clean.py 訂單.xlsx 已整理_訂單.xlsx")
        sys.exit(1)

    src_file = sys.argv[1]
    out_file = sys.argv[2] if len(sys.argv) > 2 else (
        os.path.join(
            os.path.dirname(src_file) if os.path.dirname(src_file) else '.',
            f"{os.path.splitext(os.path.basename(src_file))[0]}_已整理.xlsx"
        )
    )

    if not os.path.exists(src_file):
        print(f"錯誤: 找不到檔案: {src_file}")
        sys.exit(1)

    # 備份
    bak_file = src_file.replace('.xlsx', '_backup.xlsx')
    if not os.path.exists(bak_file):
        shutil.copy2(src_file, bak_file)
        print(f"備份已建立: {bak_file}")

    wb = openpyxl.load_workbook(src_file)
    ws = wb.active

    max_row = ws.max_row
    max_col = ws.max_column

    print(f"原始資料：{max_row} 列 x {max_col} 欄")
    print()

    # ═══════════════════════════════════════════
    # Step 1：刪除完全空白列
    # ═══════════════════════════════════════════
    print("[Step 1] 刪除空白列...")
    empty_rows = []
    for row in range(max_row, 1, -1):
        is_empty = True
        for col in range(1, max_col + 1):
            if ws.cell(row=row, column=col).value is not None:
                is_empty = False
                break
        if is_empty:
            empty_rows.append(row)

    for r in empty_rows:
        ws.delete_rows(r)

    print(f"  -> 已刪除 {len(empty_rows)} 列空白: {list(reversed(empty_rows)) if empty_rows else '無'}")
    print()

    # ═══════════════════════════════════════════
    # Step 2：子項目補上父訂單資料（含格式）
    # ═══════════════════════════════════════════
    print("[Step 2] 子項目補父訂單資料...")
    fill_cols = list(range(1, max_col + 1))

    parent = {}
    sub_count = 0
    for row in range(2, ws.max_row + 1):
        oid = ws.cell(row=row, column=1).value       # 訂單編號
        pname = ws.cell(row=row, column=31).value     # 商品名稱

        if oid is not None and str(oid).strip() != '':
            # 父訂單：記錄所有欄位的值 + 儲存格格式
            parent = {}
            for col in fill_cols:
                parent[col] = ws.cell(row=row, column=col).value
                parent[f'_fmt_{col}'] = ws.cell(row=row, column=col).number_format
        elif pname is not None and str(pname).strip() != '' and parent:
            # 子項目：補上父訂單資料
            filled = 0
            for col in fill_cols:
                cur_val = ws.cell(row=row, column=col).value
                if cur_val is None or str(cur_val).strip() == '':
                    pv = parent.get(col)
                    if pv is not None and str(pv).strip() != '':
                        ws.cell(row=row, column=col).value = pv
                        # 一併複製儲存格格式（防止日期、數字顯示不一致）
                        pf = parent.get(f'_fmt_{col}', 'General')
                        ws.cell(row=row, column=col).number_format = pf
                        filled += 1
            if filled:
                sub_count += 1
                print(f"  -> Row {row}: 子項「{str(pname).strip()[:30]}」補入 {filled} 欄（含格式）")

    print(f"  -> 共 {sub_count} 列子項目補齊")
    print()

    # ═══════════════════════════════════════════
    # Step 3：商品編號智慧校正
    # ═══════════════════════════════════════════
    # AF(第32欄=商品編號) 若非數字，且 AI(第35欄=商品型號) 為數字，則複製
    # ═══════════════════════════════════════════
    print("[Step 3] 商品編號智慧校正...")
    fix_count = 0
    for row in range(2, ws.max_row + 1):
        af_raw = ws.cell(row=row, column=32).value
        ai_raw = ws.cell(row=row, column=35).value

        af_clean = clean(af_raw)
        ai_clean = clean(ai_raw)

        if af_clean and not is_pure_digits(af_clean):
            if ai_clean and is_pure_digits(ai_clean):
                ws.cell(row=row, column=32).value = ai_raw
                # 也複製 AI 的儲存格格式到 AF
                ws.cell(row=row, column=32).number_format = ws.cell(row=row, column=35).number_format
                fix_count += 1
                print(f"  -> Row {row}: 「{af_clean}」→「{ai_clean}」（取自商品型號）")

    print(f"  -> 共 {fix_count} 列商品編號校正")
    print()

    # ═══════════════════════════════════════════
    # Step 4：刪除 AE 欄為「[促銷贈品]免運費」的列
    # ═══════════════════════════════════════════
    # AE(第31欄=商品名稱) 若包含「[促銷贈品]免運費」，則整列刪除
    # ═══════════════════════════════════════════
    print("[Step 4] 刪除 AE 欄為「[促銷贈品]免運費」的列...")
    promo_rows = []
    for row in range(ws.max_row, 1, -1):
        ae_val = ws.cell(row=row, column=31).value
        if ae_val and '[促銷贈品]免運費' in str(ae_val):
            promo_rows.append(row)

    for r in promo_rows:
        ws.delete_rows(r)

    print(f"  -> 已刪除 {len(promo_rows)} 列「[促銷贈品]免運費」: {list(reversed(promo_rows)) if promo_rows else '無'}")
    print()

    # ═══════════════════════════════════════════
    # 另存新檔
    # ═══════════════════════════════════════════
    wb.save(out_file)
    print("=" * 50)
    print(f"整理完成！")
    print(f"  輸出: {out_file}")
    print(f"  最終: {ws.max_row} 列 x {ws.max_column} 欄（{ws.max_row - 1} 筆資料）")


if __name__ == '__main__':
    main()
